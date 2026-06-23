"""
Apple Procurement Monitor
Crawls muasamcong.mpi.gov.vn for Apple/IT-related bids, deduplicates
against Google Sheet, analyzes with Claude, and emails new bids.

Usage:
  python apple_monitor.py          — full run
  python apple_monitor.py test     — send a test email with 1 dummy record
"""

import io
import re
import ssl
import sys
import time
import json
import random
import smtplib
import urllib3
import urllib.parse
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formataddr

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import requests
import gspread
from google import genai
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build as _drive_build
from googleapiclient.http import MediaIoBaseUpload

urllib3.disable_warnings()

# Strip XML-illegal control characters that openpyxl rejects
_ILLEGAL_CHARS   = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F￾￿]")
# Characters that trigger formula/DDE injection in Excel when leading a cell
_FORMULA_LEADERS = frozenset(("=", "+", "-", "@", "\t", "\r"))

def _clean(val):
    if isinstance(val, str):
        val = _ILLEGAL_CHARS.sub("", val)
        if val and val[0] in _FORMULA_LEADERS:
            val = " " + val  # prepend space — Excel treats as plain text
        return val
    return val

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from apple_monitor_config import (
    INVEST_FIELDS,
    GMAIL_SENDER, GMAIL_APP_PASSWORD, EMAIL_RECIPIENTS,
    GOOGLE_SHEET_ID, GOOGLE_CREDENTIALS_DICT,
    GEMINI_API_KEY,
    API_TOKEN, API_COOKIES,
)

# ── Constants ──────────────────────────────────────────────────
BASE_URL  = "https://muasamcong.mpi.gov.vn"
API_PATH  = "/o/egp-portal-contractor-selection-v2/services/smart/search"
PAGE_SIZE = 50
DELAY_MIN = 1.5
DELAY_MAX = 3.0

SHEET_NAME = "Bids"
SHEET_COLS = [
    "notifyId", "keyword", "notifyNo", "bid_name",
    "investorName", "investorCode", "prov_name",
    "publicDate", "bidCloseDate", "priceInit",
    "bidForm", "bidMode", "status",
    "analysis", "crawled_at", "source_url",
    "winner", "winner_price", "goods_url",
]

STATUS_MAP = {
    "IS_PUBLISH":  "Open",
    "OPEN_BID":    "Bidding Open",
    "OPEN_DXTC":   "Under Evaluation",
    "OPEN_DXKT":   "Technical Evaluation",
    "PUB_KQLCNT":  "Result Published",
    "PUB_MT":      "Invitation Published",
    "CANCEL_BID":  "Cancelled",
    "CANCELED":    "Cancelled",
    "IS_CANCEL":   "Cancelled",
    "NEW":         "New",
    "INIT_MT":     "Initializing",
    "1":           "Open",
    "3":           "Closed",
}


# ── SSL adapter (weak DH key on government server) ────────────

class LegacySSLAdapter(requests.adapters.HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.create_default_context()
        ctx.set_ciphers("DEFAULT:@SECLEVEL=0")
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        kwargs["ssl_context"] = ctx
        return super().init_poolmanager(*args, **kwargs)


HEADERS = {
    "User-Agent":   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept":       "application/json, text/plain, */*",
    "Content-Type": "application/json",
    "Origin":       BASE_URL,
    "Referer":      f"{BASE_URL}/web/guest/contractor-selection?render=index",
}


# ── Crawler ────────────────────────────────────────────────────

def make_session():
    s = requests.Session()
    s.mount("https://", LegacySSLAdapter())
    s.cookies.update(API_COOKIES)
    return s


def _build_payload(page_number, page_size, invest_fields=None, since_date=None):
    """
    since_date: "YYYY-MM-DD" — chỉ lấy bids publicDate >= since_date.
    Dùng cho daily incremental crawl để tránh crawl toàn bộ lịch sử.
    """
    if invest_fields is None:
        invest_fields = INVEST_FIELDS
    filters = [
        {"fieldName": "investField", "searchType": "in",     "fieldValues": invest_fields},
        {"fieldName": "type",        "searchType": "in",     "fieldValues": ["es-notify-contractor"]},
        {"fieldName": "caseKHKQ",    "searchType": "not_in", "fieldValues": ["1"]},
    ]
    if since_date:
        filters.append({"fieldName": "publicDate", "searchType": "gte", "fieldValues": [since_date]})
    return [{
        "pageSize": page_size,
        "pageNumber": str(page_number),
        "query": [{
            "index": "es-contractor-selection",
            "matchType": "all-1",
            "matchFields": ["notifyNo", "bidName"],
            "filters": filters,
        }],
    }]


def _fetch_page(session, page_number, invest_fields=None, page_size=PAGE_SIZE, since_date=None):
    url = f"{BASE_URL}{API_PATH}?token={API_TOKEN}"
    payload = _build_payload(page_number, page_size, invest_fields, since_date)
    for attempt in range(3):
        try:
            resp = session.post(url, headers=HEADERS, json=payload, verify=False, timeout=20)
            if resp.status_code == 401:
                print("  [401] Token expired — refresh API_TOKEN in apple_monitor_config.py")
                return None
            if resp.status_code in (403, 429):
                wait = 30 * (attempt + 1)
                print(f"  [{resp.status_code}] rate-limited — waiting {wait}s")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            data = resp.json()
            return data[0] if isinstance(data, list) else data
        except requests.RequestException as e:
            print(f"  [retry {attempt+1}/3] {e}")
            time.sleep(5 * (attempt + 1))
    return None


def crawl_categories(session, invest_fields=None, since_date=None, max_pages=40):
    """
    Crawl by investField category instead of keyword.
    since_date: "YYYY-MM-DD" — filter bids publicDate >= since_date (nếu portal hỗ trợ).
    max_pages: giới hạn số trang để tránh timeout và IP block. Default=40 (~2000 bids).
    """
    if invest_fields is None:
        invest_fields = INVEST_FIELDS
    label = "+".join(invest_fields)
    date_note = f" since {since_date}" if since_date else ""

    data = _fetch_page(session, 0, invest_fields, since_date=since_date)
    if not data:
        return []

    page_obj      = data.get("page", data)
    total_pages   = min(page_obj.get("totalPages", 1), max_pages)
    total_records = page_obj.get("totalElements", 0)
    capped = " (capped)" if page_obj.get("totalPages", 1) > max_pages else ""
    print(f"    [{label}]{date_note}: {total_records:,} records — crawling {total_pages} pages{capped}")

    all_items = list(data.get("content") or data.get("page", {}).get("content", []))

    for page_num in range(1, total_pages):
        page_data = _fetch_page(session, page_num, invest_fields, since_date=since_date)
        if not page_data:
            break
        content = page_data.get("content") or page_data.get("page", {}).get("content", [])
        all_items.extend(content)
        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    return all_items


def _build_source_url(item):
    uid       = item.get("id") or item.get("notifyId", "")
    step_code = item.get("stepCode", "")
    step      = step_code.split("-")[-1] if step_code else "tbmt"

    def _p(v):
        return v if v not in (None, "", []) else "undefined"

    params = (
        "p_p_id=egpportalcontractorselectionv2_WAR_egpportalcontractorselectionv2"
        "&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view"
        "&_egpportalcontractorselectionv2_WAR_egpportalcontractorselectionv2_render=detail-v2"
        f"&type=es-notify-contractor"
        f"&stepCode={_p(step_code)}"
        f"&id={_p(uid)}"
        f"&notifyId={_p(uid)}"
        f"&inputResultId={_p(item.get('inputResultId'))}"
        f"&bidOpenId={_p(item.get('bidOpenId'))}&techReqId=undefined"
        f"&bidPreNotifyResultId=undefined&bidPreOpenId=undefined"
        f"&processApply={_p(item.get('processApply'))}"
        f"&bidMode={_p(item.get('bidMode'))}"
        f"&notifyNo={_p(item.get('notifyNo'))}"
        f"&planNo={_p(item.get('planNo'))}"
        f"&pno=undefined"
        f"&step={step}"
        f"&isInternet={_p(item.get('isInternet'))}"
        f"&caseKHKQ={_p(item.get('caseKHKQ'))}"
        f"&bidForm={_p(item.get('bidForm'))}"
    )
    return f"{BASE_URL}/web/guest/contractor-selection?{params}"


def flatten(item, keyword="HH"):
    locs      = item.get("locations") or []
    first_loc = locs[0] if locs else {}
    bid_name  = item.get("bidName", "")
    if isinstance(bid_name, list):
        bid_name = " | ".join(str(v) for v in bid_name if v)
    price = item.get("priceInit") or (item.get("bidPrice") or [""])[0] if isinstance(item.get("bidPrice"), list) else item.get("bidPrice", "")

    winner_raw = item.get("winningContractorName", "")
    winner = " | ".join(str(v) for v in winner_raw if v) if isinstance(winner_raw, list) else str(winner_raw or "")

    price_raw = item.get("bidWinningPrice", "")
    winner_price = price_raw[0] if isinstance(price_raw, list) and price_raw else (price_raw or "")

    return {
        "notifyId":     item.get("notifyId") or item.get("id", ""),
        "keyword":      keyword,
        "notifyNo":     item.get("notifyNo", ""),
        "bid_name":     bid_name,
        "investorName": item.get("investorName", ""),
        "investorCode": item.get("investorCode", ""),
        "prov_name":    first_loc.get("provName", ""),
        "publicDate":   (item.get("publicDate") or "")[:10],
        "bidCloseDate": (item.get("bidCloseDate") or "")[:10],
        "priceInit":    price,
        "bidForm":      item.get("bidForm", ""),
        "bidMode":      item.get("bidMode", ""),
        "status":       item.get("status", ""),
        "analysis":     "",
        "crawled_at":   datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source_url":   _build_source_url(item),
        "winner":       winner,
        "winner_price": winner_price,
        "goods_url":    "",
    }


# ── Google Sheets ──────────────────────────────────────────────

def connect_sheet():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(GOOGLE_CREDENTIALS_DICT, scopes=scopes)
    gc    = gspread.authorize(creds)
    sh    = gc.open_by_key(GOOGLE_SHEET_ID)
    try:
        ws = sh.worksheet(SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=SHEET_NAME, rows=50000, cols=len(SHEET_COLS))
        ws.append_row(SHEET_COLS)
        ws.format("1", {"textFormat": {"bold": True}})
    return ws


def get_existing_ids(ws):
    col_idx = SHEET_COLS.index("notifyId") + 1
    vals    = ws.col_values(col_idx)
    return set(vals[1:])  # skip header row


def append_to_sheet(ws, records):
    rows = [[str(r.get(c, "")) for c in SHEET_COLS] for r in records]
    ws.append_rows(rows, value_input_option="RAW")


# ── Claude analysis ────────────────────────────────────────────

def _days_left(date_str):
    try:
        dt = datetime.fromisoformat(str(date_str)[:10])
        return (dt.date() - date.today()).days
    except Exception:
        return None


_QUOTA_EXCEEDED = "__QUOTA_EXCEEDED__"


def analyze_bid(client, record):
    price = record.get("priceInit", "")
    try:
        price_str = f"{float(price)/1e9:.2f} tỷ VND"
    except Exception:
        price_str = str(price) if price else "N/A"

    days = _days_left(record.get("bidCloseDate"))
    deadline_str = f"{days} days left ({record.get('bidCloseDate','')})" if days is not None else record.get("bidCloseDate", "N/A")

    prompt = f"""Analyze this Vietnamese public procurement bid briefly. Reply in English only.

Bid name: {record.get('bid_name', '')}
Buyer: {record.get('investorName', '')}
Province: {record.get('prov_name', '')}
Value: {price_str}
Deadline: {deadline_str}
Search keyword: {record.get('keyword', '')}

Reply in exactly 4 lines, no extra explanation:
Product: [specific Apple/IT device type]
Value: [summary]
Deadline: [X days left — urgent/normal/plenty of time]
Priority: [High/Medium/Low] — [one-line reason]"""

    for attempt in range(4):
        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash-lite",
                contents=prompt,
            )
            return response.text.strip()
        except Exception as e:
            err = str(e)
            if "429" in err or "RESOURCE_EXHAUSTED" in err or "quota" in err.lower():
                return _QUOTA_EXCEEDED
            if ("503" in err or "500" in err or "UNAVAILABLE" in err) and attempt < 3:
                time.sleep(5 * (attempt + 1))  # 5s, 10s, 15s
                continue
            return f"(analysis error: {e})"


# ── Email ──────────────────────────────────────────────────────

def _fmt_price(val):
    try:
        return f"{float(val)/1e9:.2f}B VND".replace(".", ",")
    except Exception:
        return str(val) if val else "—"


def _fmt_date(val):
    if not val:
        return "—"
    try:
        dt   = datetime.fromisoformat(str(val)[:10])
        days = (dt.date() - date.today()).days
        warn = " (!)" if 0 <= days <= 5 else ""
        return f"{dt.strftime('%d/%m/%Y')}{warn} ({days}d left)"
    except Exception:
        return str(val)[:10]


def _sort_by_value(records):
    def _val(r):
        try:
            return float(r.get("priceInit") or 0)
        except Exception:
            return 0.0
    return sorted(records, key=_val, reverse=True)


def _build_html(records, note_html=""):
    now   = datetime.now().strftime("%Y-%m-%d %H:%M")
    total = len(records)

    active = _sort_by_value(
        [r for r in records if (_days_left(r.get("bidCloseDate")) or -1) >= 0]
    )

    if active:
        cards = ""
        for i, r in enumerate(active, 1):
            analysis = (r.get("analysis") or "").replace("\n", "<br>")
            kw = r.get("keyword", "")
            cards += f"""
<div style="border:1px solid #dadce0;border-radius:8px;padding:16px;margin-bottom:10px;background:#fff">
  <div style="display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:10px">
    <div style="font-size:14px;font-weight:600;color:#202124;flex:1">{i}. {r.get('bid_name','—')}</div>
    <span style="margin-left:10px;flex-shrink:0;padding:2px 8px;background:#e8f0fe;color:#1a73e8;
                 font-size:11px;font-weight:600;border-radius:4px;white-space:nowrap">{kw}</span>
  </div>
  <table style="font-size:13px;color:#555;border-collapse:collapse;width:100%">
    <tr><td style="width:100px;color:#888;padding:2px 0">Buyer:</td>
        <td><b>{r.get('investorName','—')}</b></td></tr>
    <tr><td style="color:#888;padding:2px 0">Province:</td>
        <td>{r.get('prov_name','—')}</td></tr>
    <tr><td style="color:#888;padding:2px 0">Value:</td>
        <td><b style="color:#1a73e8">{_fmt_price(r.get('priceInit',''))}</b></td></tr>
    <tr><td style="color:#888;padding:2px 0">Deadline:</td>
        <td>{_fmt_date(r.get('bidCloseDate',''))}</td></tr>
    <tr><td style="color:#888;padding:2px 0">Ref no.:</td>
        <td style="font-family:monospace;font-size:12px">{r.get('notifyNo','—')}</td></tr>
  </table>
  <div style="margin-top:10px;padding:10px 12px;background:#f8f9fa;border-left:3px solid #1a73e8;
              font-size:13px;color:#333;line-height:1.6;border-radius:0 4px 4px 0">
    {analysis}
  </div>
  <div style="margin-top:10px">
    <a href="{r.get('source_url','')}" target="_blank"
       style="display:inline-block;padding:7px 16px;background:#1a73e8;color:#fff;
              font-size:12px;font-weight:600;border-radius:5px;text-decoration:none">
      View on Portal →
    </a>
  </div>
</div>"""

        sections = f'<div style="margin-top:16px">{cards}</div>'
    else:
        sections = """
<div style="text-align:center;padding:32px;color:#888;font-size:14px">
  No active bids at this time. See the attached Excel for all new records.
</div>"""

    security_note = """
<div style="background:#fff8e1;border:1px solid #ffe082;border-radius:8px;
            padding:12px 16px;margin-bottom:20px;font-size:13px;color:#5d4037">
  <b>Security reminder:</b> A detailed Excel report is attached to this email.
  Before opening, verify the file extension is <b>.xlsx</b>.
  Do <b>not</b> open files with unexpected extensions
  (<code>.exe</code>, <code>.zip</code>, <code>.bat</code>, etc.) — they may contain malware.
</div>"""

    return f"""<!DOCTYPE html>
<html><body style="font-family:'Google Sans',Arial,sans-serif;max-width:680px;
                   margin:0 auto;padding:0;background:#f1f3f4;color:#202124">
<div style="background:#1a73e8;padding:22px 28px;border-radius:10px 10px 0 0">
  <div style="display:flex;align-items:center;gap:14px">
    <div style="width:38px;height:38px;background:rgba(255,255,255,0.15);border-radius:9px;
                display:inline-flex;align-items:center;justify-content:center;
                font-size:22px;color:#fff;flex-shrink:0;font-family:-apple-system,sans-serif">
      &#xF8FF;
    </div>
    <div>
      <div style="font-size:18px;font-weight:700;color:#fff;letter-spacing:-.3px;line-height:1.2">
        Procurement Monitor
      </div>
      <div style="color:#c5d9f7;font-size:12px;margin-top:3px">
        {total} new bid(s) detected ({len(active)} active) &mdash; {now}
      </div>
    </div>
  </div>
</div>
<div style="background:#fff;padding:24px 28px;border:1px solid #dadce0;
            border-top:none;border-radius:0 0 10px 10px">
  {note_html}
  {security_note}
  {sections}
  <div style="margin-top:28px;padding-top:16px;border-top:1px solid #e8eaed;
              font-size:12px;color:#aaa;text-align:center">
    Showing active bids only &mdash; full report in attached <b>.xlsx</b> file
    &nbsp;|&nbsp; Source: <a href="https://muasamcong.mpi.gov.vn" style="color:#1a73e8">muasamcong.mpi.gov.vn</a>
  </div>
</div>
</body></html>"""


COLUMNS = [
    ("No.",              5),
    ("Ref No.",          18),
    ("Bid Name",         45),
    ("Buyer",            30),
    ("Province",         16),
    ("Value (B VND)",    14),
    ("Created Date",     13),
    ("Deadline",         13),
    ("Days Left",        10),
    ("Status",           14),
    ("Winner",           35),
    ("Award (B VND)",    14),
    ("Bid Form",         16),
    ("Keyword",          16),
    ("AI Analysis",      45),
    ("Link",             20),
]


def _write_sheet(ws, records):
    header_fill = PatternFill("solid", fgColor="1a73e8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="DADCE0")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    urgent_fill = PatternFill("solid", fgColor="FDECEA")
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill    = PatternFill("solid", fgColor="F8F9FA")

    ws.row_dimensions[1].height = 28
    for col_idx, (label, width) in enumerate(COLUMNS, 1):
        cell           = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    for row_idx, r in enumerate(records, 2):
        days = _days_left(r.get("bidCloseDate"))
        try:
            value_b = round(float(r.get("priceInit", 0)) / 1e9, 2)
        except Exception:
            value_b = ""

        created_str = ""
        if r.get("publicDate"):
            try:
                dt = datetime.fromisoformat(str(r["publicDate"])[:10])
                created_str = dt.strftime("%d/%m/%Y")
            except Exception:
                created_str = str(r["publicDate"])[:10]

        deadline_str = ""
        if r.get("bidCloseDate"):
            try:
                dt = datetime.fromisoformat(str(r["bidCloseDate"])[:10])
                deadline_str = dt.strftime("%d/%m/%Y")
            except Exception:
                deadline_str = str(r["bidCloseDate"])[:10]

        try:
            award_b = round(float(r.get("winner_price", 0) or 0) / 1e9, 2) if r.get("winner_price") else ""
        except Exception:
            award_b = ""

        status_display = STATUS_MAP.get(r.get("status", ""), r.get("status", ""))

        row_data = [
            row_idx - 1,                        # 1  No.
            r.get("notifyNo", ""),              # 2  Ref No.
            r.get("bid_name", ""),              # 3  Bid Name
            r.get("investorName", ""),          # 4  Buyer
            r.get("prov_name", ""),             # 5  Province
            value_b,                            # 6  Value (B VND)
            created_str,                        # 7  Created Date
            deadline_str,                       # 8  Deadline
            days if days is not None else "",   # 9  Days Left
            status_display,                     # 10 Status
            r.get("winner", ""),                # 11 Winner
            award_b,                            # 12 Award (B VND)
            r.get("bidForm", ""),               # 13 Bid Form
            r.get("keyword", ""),               # 14 Keyword
            r.get("analysis", ""),              # 15 AI Analysis
            r.get("source_url", ""),            # 16 Link
        ]

        fill = urgent_fill if (days is not None and 0 <= days <= 5) \
               else (alt_fill if row_idx % 2 == 0 else normal_fill)
        ws.row_dimensions[row_idx].height = 60

        from openpyxl.worksheet.hyperlink import Hyperlink
        for col_idx, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=_clean(value))
            cell.border    = border
            cell.fill      = fill
            cell.alignment = center if col_idx in (1, 6, 7, 8, 9, 12) else left
            if col_idx == 6 and isinstance(value, float):
                cell.number_format = '#,##0.00'
            if col_idx == 12 and isinstance(value, float):
                cell.number_format = '#,##0.00'
            if col_idx == 9 and isinstance(value, int):
                cell.font = Font(bold=True, color="C62828" if 0 <= value <= 5 else "202124")
            if col_idx == 16 and value:
                cell.value     = "View on Portal"
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=value)
                cell.font      = Font(color="1155CC", underline="single")

    # Summary row
    last = len(records) + 2
    sum_cell = ws.cell(row=last, column=1, value="Total")
    sum_cell.font      = Font(bold=True)
    sum_cell.alignment = center
    val_cell = ws.cell(row=last, column=6, value=f"=SUM(F2:F{last-1})")
    val_cell.number_format = '#,##0.00'
    val_cell.font          = Font(bold=True)


def _build_excel(new_records, all_records=None) -> bytes:
    if all_records is None:
        all_records = new_records
    # Active tab = ALL active bids from full sheet (not just new ones)
    active = [r for r in all_records if (_days_left(r.get("bidCloseDate")) or -1) >= 0]

    wb = openpyxl.Workbook()

    def _sort_active(r):
        try: v = float(r.get("priceInit", 0) or 0)
        except: v = 0
        days = _days_left(r.get("bidCloseDate"))
        return (-v, days if days is not None else 9999)

    def _sort_val(r):
        try: return float(r.get("priceInit", 0) or 0)
        except: return 0

    # Tab 1 — All currently active bids: Value desc, then Days Left asc
    ws_active       = wb.active
    ws_active.title = f"Active ({len(active)})"
    _write_sheet(ws_active, sorted(active, key=_sort_active))

    # Tab 2 — Full database export from Google Sheet: Value desc
    ws_all       = wb.create_sheet(title=f"All Records ({len(all_records)})")
    _write_sheet(ws_all, sorted(all_records, key=_sort_val, reverse=True))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_email(records, all_records=None, recipients=None, note_html="", subject=None):
    to = recipients if recipients is not None else EMAIL_RECIPIENTS
    html     = _build_html(records, note_html=note_html)
    xlsx     = _build_excel(records, all_records)
    if subject is None:
        subject = f"[Procurement] {len(records)} new bid(s) — {datetime.now().strftime('%Y-%m-%d')}"
    filename = f"bids_{datetime.now().strftime('%Y%m%d')}.xlsx"

    msg            = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = formataddr((" Procurement Monitor", GMAIL_SENDER))
    msg["To"]      = ", ".join(to)

    msg.attach(MIMEText(html, "html", "utf-8"))

    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(xlsx)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as srv:
        srv.login(GMAIL_SENDER, GMAIL_APP_PASSWORD)
        srv.sendmail(GMAIL_SENDER, to, msg.as_string())

    print(f"  Email sent → {', '.join(to)} (attachment: {filename})")


# ── Status refresh ────────────────────────────────────────────

# ── Goods list helpers ─────────────────────────────────────────

_ENDPOINT_LDT = "/o/egp-portal-contractor-selection-v2/services/expose/contractor-input-result/get"
_ENDPOINT_VK  = "/o/egp-portal-contractor-selection-v2/services/expose/kqlcnt/bid-notify-contractor-out/get-by-id"


def _url_param(source_url, param):
    val = urllib.parse.parse_qs(urllib.parse.urlparse(source_url).query).get(param, [""])[0]
    return "" if val == "undefined" else val


def _fetch_goods_items(session, input_result_id, process_apply="LDT"):
    endpoint = _ENDPOINT_VK if process_apply == "VK" else _ENDPOINT_LDT
    url = f"{BASE_URL}{endpoint}?token={API_TOKEN}"
    try:
        resp = session.post(url, json={"id": input_result_id}, verify=False, timeout=20)
        if resp.status_code != 200:
            return []
        dto  = resp.json().get("bideContractorInputResultDTO", {})
        lots = dto.get("lotResultDTO", [])
        if not lots:
            return []
        gs_raw = lots[0].get("goodsList") or ""
        if not gs_raw:
            return []
        gs    = json.loads(gs_raw)
        table = gs[0].get("formValue", {}).get("lotContent", {}).get("Table", [])
        return table if isinstance(table, list) else []
    except Exception:
        return []


def _create_goods_excel(items, bid_name, notify_no, winner=""):
    from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlign
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Danh muc hang hoa"

    for text in [f"Gói thầu: {bid_name}", f"Số thông báo: {notify_no}",
                 f"Nhà thầu trúng: {winner}" if winner else None, None]:
        if text is not None:
            ws.append([text])
        else:
            ws.append([])

    headers = ["STT", "Tên hàng hóa", "Nhãn hiệu", "Model / Mã hàng",
               "Số lượng", "ĐVT", "Xuất xứ", "Đơn giá (VND)", "Thành tiền (VND)", "Thông số kỹ thuật"]
    ws.append(headers)
    hrow = ws.max_row
    hfill = XFill(fill_type="solid", fgColor="1a73e8")
    for cell in ws[hrow]:
        cell.font      = XFont(bold=True, color="FFFFFF")
        cell.fill      = hfill
        cell.alignment = XAlign(wrap_text=True, horizontal="center", vertical="center")

    for i, item in enumerate(items, 1):
        brand = item.get("labelGood") or item.get("lableGood") or item.get("manufacturer", "")
        feat  = (item.get("feature") or "")[:1000]
        ws.append([
            i,
            item.get("name", ""),
            brand,
            item.get("codeGood") or item.get("model", ""),
            item.get("qty", ""),
            item.get("uom", ""),
            item.get("origin", ""),
            item.get("bidPrice") or "",
            item.get("amount")   or "",
            feat,
        ])
        ws.row_dimensions[ws.max_row].height = 40

    for col_idx, width in enumerate([5, 40, 18, 20, 10, 8, 12, 18, 18, 70], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in (8, 9):
        for row in ws.iter_rows(min_row=hrow+1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_to_drive(xlsx_bytes, filename):
    creds   = Credentials.from_service_account_info(
        GOOGLE_CREDENTIALS_DICT,
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    service = _drive_build("drive", "v3", credentials=creds, cache_discovery=False)
    media   = MediaIoBaseUpload(
        io.BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    f = service.files().create(
        body={"name": filename},
        media_body=media,
        fields="id"
    ).execute()
    service.permissions().create(
        fileId=f["id"],
        body={"type": "anyone", "role": "reader"}
    ).execute()
    return f"https://drive.google.com/file/d/{f['id']}/view"


def refresh_goods_urls(ws, session):
    """Upload goods Excel to Drive for PUB_KQLCNT bids that have inputResultId but no goods_url."""
    rows = ws.get_all_records(head=1)
    to_process = [
        (i + 2, r) for i, r in enumerate(rows)
        if r.get("status") == "PUB_KQLCNT"
        and not str(r.get("goods_url", "")).strip()
        and _url_param(r.get("source_url", ""), "inputResultId")
    ]

    if not to_process:
        print("  No goods URLs to refresh.")
        return

    print(f"  {len(to_process)} bids need goods upload...")
    gc      = get_column_letter(SHEET_COLS.index("goods_url") + 1)
    updates = []
    done    = 0

    for sheet_row, record in to_process:
        source_url      = record.get("source_url", "")
        input_result_id = _url_param(source_url, "inputResultId")
        process_apply   = _url_param(source_url, "processApply") or "LDT"

        # Fallback: if inputResultId missing, re-fetch bid by notifyNo to get full detail URL
        if not input_result_id and record.get("notifyNo"):
            try:
                api_url = f"{BASE_URL}{API_PATH}?token={API_TOKEN}"
                payload = [{"pageSize": 1, "pageNumber": 0, "query": [{
                    "index": "es-contractor-selection", "keyWord": record["notifyNo"],
                    "matchType": "all-1", "matchFields": ["notifyNo"],
                    "filters": [{"fieldName": "type", "searchType": "in",
                                 "fieldValues": ["es-notify-contractor"]}],
                }]}]
                resp  = session.post(api_url, headers=HEADERS, json=payload, verify=False, timeout=15)
                items_api = (resp.json()[0] if isinstance(resp.json(), list) else resp.json()).get("page", {}).get("content", [])
                if items_api:
                    input_result_id = items_api[0].get("inputResultId") or ""
                    process_apply   = items_api[0].get("processApply") or "LDT"
            except Exception:
                pass

        if not input_result_id:
            continue

        items = _fetch_goods_items(session, input_result_id, process_apply)
        if not items:
            continue

        notify_no = record.get("notifyNo", "unknown")
        safe_name = re.sub(r"[^\w\-]", "_", notify_no)
        filename  = f"HangHoa_{safe_name}.xlsx"

        try:
            xlsx_bytes = _create_goods_excel(
                items,
                bid_name  = record.get("bid_name", ""),
                notify_no = notify_no,
                winner    = record.get("winner", ""),
            )
            drive_url = _upload_to_drive(xlsx_bytes, filename)
            updates.append({"range": f"{gc}{sheet_row}", "values": [[drive_url]]})
            done += 1
            print(f"    {notify_no}: {len(items)} items → {drive_url[:60]}...")
        except Exception as e:
            print(f"    [err] {notify_no}: {e}")

        time.sleep(random.uniform(1.0, 1.5))

        if len(updates) >= 50:
            ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates})
            updates = []

    if updates:
        ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates})
    print(f"  Goods URLs uploaded: {done}")


_STALE_STATUSES = {"IS_PUBLISH", "1", "OPEN_BID", "NEW", "INIT_MT", "OPEN_DXTC", "OPEN_DXKT"}


_MAX_REFRESH_PER_RUN = 200


def refresh_recently_closed(ws, session, lookback_days=None):
    """
    Re-fetch status + winner + price + source_url for closed bids with stale status.
    Processes the _MAX_REFRESH_PER_RUN most-recently-closed bids per run so each
    run stays under ~2 minutes. Backlog clears gradually across daily runs.
    """
    rows      = ws.get_all_records(head=1)
    today_str = str(date.today())

    candidates = [
        (i + 2, r)
        for i, r in enumerate(rows)
        if r.get("bidCloseDate", "") < today_str
        and r.get("status", "") in _STALE_STATUSES
        and r.get("notifyNo", "").strip()
    ]

    # Prioritize most-recently-closed; cap to avoid long runs
    candidates.sort(key=lambda x: x[1].get("bidCloseDate", ""), reverse=True)
    to_refresh = candidates[:_MAX_REFRESH_PER_RUN]

    if not to_refresh:
        print("  No closed bids with stale status.")
        return

    total_stale = len(candidates)
    print(f"  {total_stale} stale bids total — refreshing {len(to_refresh)} most recent...")

    sc  = get_column_letter(SHEET_COLS.index("status") + 1)
    wc  = get_column_letter(SHEET_COLS.index("winner") + 1)
    pc  = get_column_letter(SHEET_COLS.index("winner_price") + 1)
    uc  = get_column_letter(SHEET_COLS.index("source_url") + 1)

    url     = f"{BASE_URL}{API_PATH}?token={API_TOKEN}"
    updates = []
    done    = 0

    for sheet_row, record in to_refresh:
        notify_no = record["notifyNo"]
        payload = [{
            "pageSize": 1, "pageNumber": 0,
            "query": [{
                "index": "es-contractor-selection",
                "keyWord": notify_no,
                "matchType": "all-1",
                "matchFields": ["notifyNo"],
                "filters": [{"fieldName": "type", "searchType": "in",
                             "fieldValues": ["es-notify-contractor"]}],
            }],
        }]
        try:
            resp = session.post(url, headers=HEADERS, json=payload, verify=False, timeout=15)
            data  = resp.json()
            items = (data[0] if isinstance(data, list) else data).get("page", {}).get("content", [])
            if not items:
                continue
            it         = items[0]
            new_status = it.get("status", "") or record.get("status", "")

            winner_raw = it.get("winningContractorName", "")
            winner     = (" | ".join(str(v) for v in winner_raw if v)
                          if isinstance(winner_raw, list) else str(winner_raw or ""))
            price_raw  = it.get("bidWinningPrice", "")
            price      = (str(price_raw[0]) if isinstance(price_raw, list) and price_raw
                          else str(price_raw or ""))

            updates.append({"range": f"{sc}{sheet_row}", "values": [[new_status]]})
            if winner and not record.get("winner", "").strip():
                updates.append({"range": f"{wc}{sheet_row}", "values": [[winner]]})
                updates.append({"range": f"{pc}{sheet_row}", "values": [[price]]})

            # Rebuild source_url if the refreshed item has richer IDs (inputResultId, bidOpenId)
            new_url = _build_source_url(it)
            old_url = record.get("source_url", "")
            if new_url != old_url and "inputResultId=undefined" not in new_url:
                updates.append({"range": f"{uc}{sheet_row}", "values": [[new_url]]})

            done += 1
        except Exception as e:
            print(f"  [err] {notify_no}: {e}")

        if len(updates) >= 300:
            ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates})
            updates = []

        time.sleep(random.uniform(0.3, 0.6))

    if updates:
        ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates})

    print(f"  Refreshed {done} closed bids")


# ── Main ───────────────────────────────────────────────────────

def run(do_send_email=True):
    print(f"=== Apple Monitor — {datetime.now().strftime('%Y-%m-%d %H:%M')} ===\n")

    print("Connecting to Google Sheet...")
    ws           = connect_sheet()
    existing_ids = get_existing_ids(ws)
    print(f"  {len(existing_ids):,} existing records in sheet\n")

    session  = make_session()
    all_new  = []
    seen_ids = set(existing_ids)

    # max_pages=40 → tối đa 2000 bids/run, ~90s, an toàn cho GitHub Actions timeout 30min
    print(f"Crawling categories: {INVEST_FIELDS} (max 40 pages) ...")
    items = crawl_categories(session, INVEST_FIELDS, max_pages=40)
    for item in items:
        invest_field = item.get("investField", "HH")
        if isinstance(invest_field, list):
            invest_field = invest_field[0] if invest_field else "HH"
        record = flatten(item, invest_field)
        nid    = record["notifyId"]
        if nid and nid not in seen_ids:
            all_new.append(record)
            seen_ids.add(nid)

    print(f"\n→ {len(all_new)} new records found\n")

    print("Refreshing recently-closed bids...")
    refresh_recently_closed(ws, session)
    print()

    if not all_new:
        print("Nothing new — no email sent.")
        return True

    # Filter: chỉ Gemini-analyze bids IT-relevant (HH+HON_HOP crawl lấy tất cả hàng hóa,
    # cần loại bỏ thuốc, thực phẩm, vật tư y tế, xăng dầu, v.v.)
    TECH_KW = [
        "máy tính", "laptop", "máy tính xách tay", "máy vi tính",
        "máy tính bảng", "tablet", "ipad", "iphone", "macbook", "apple",
        "imac", "mac mini", "mac pro", "mac studio",
        "điện thoại thông minh", "smartphone", "điện thoại di động",
        "thiết bị cntt", "thiết bị công nghệ thông tin", "thiết bị it",
        "máy tính để bàn", "all-in-one", "workstation", "máy trạm",
        "server", "máy chủ", "switch", "router", "thiết bị mạng",
        "máy in", "máy scan", "scanner", "màn hình", "monitor",
        "camera", "ups", "storage", "lưu trữ",
    ]
    def _is_tech(bid_name):
        n = bid_name.lower()
        return any(k in n for k in TECH_KW)

    active_all  = [r for r in all_new if (_days_left(r.get("bidCloseDate")) or -1) >= 0]
    to_analyze  = [r for r in active_all if _is_tech(r.get("bid_name", ""))]
    skipped_exp = len(all_new) - len(active_all)
    skipped_non = len(active_all) - len(to_analyze)
    print(f"Analyzing with Gemini: {len(to_analyze)} IT-relevant active bids "
          f"(skipped {skipped_exp} expired, {skipped_non} non-IT)")
    client = genai.Client(api_key=GEMINI_API_KEY)
    quota_hit = False
    for i, record in enumerate(to_analyze):
        print(f"  [{i+1}/{len(to_analyze)}] {record['bid_name'][:70]}")
        if quota_hit:
            record["analysis"] = "(quota exceeded — re-run tomorrow)"
            continue
        result = analyze_bid(client, record)
        if result == _QUOTA_EXCEEDED:
            quota_hit = True
            print(f"  [!] Gemini daily quota reached at record {i+1} — remaining saved without analysis")
            record["analysis"] = "(quota exceeded — re-run tomorrow)"
        else:
            record["analysis"] = result
        time.sleep(0.5)
    if quota_hit:
        print("  Note: Quota exhausted. Pipeline continues — email & sheet will be sent with partial analysis.")

    print("\nWriting to Google Sheet...")
    append_to_sheet(ws, all_new)

    print("Fetching full sheet for Excel export...")
    all_rows = ws.get_all_records(head=1)
    all_sheet = [{col: str(row.get(col, "")) for col in SHEET_COLS} for row in all_rows]
    print(f"  {len(all_sheet):,} total records in sheet")

    if do_send_email:
        print("Sending email...")
        send_email(all_new, all_sheet)
    else:
        print("Crawl-only mode — skipping email.")

    print(f"\n✓ Done — {len(all_new)} new records processed.")


def run_test():
    """Send a test email to minh_dao only, with 3 dummy records sorted by value."""
    dummy = [
        {
            "notifyId":    "TEST-001",
            "keyword":     "macbook",
            "notifyNo":    "TEST20260522-A",
            "bid_name":    "[TEST] Mua sắm MacBook Pro M4 cho Bộ Khoa học và Công nghệ",
            "investorName": "Bộ Khoa học và Công nghệ",
            "investorCode": "vn0000000001",
            "prov_name":   "Hà Nội",
            "publicDate":  date.today().isoformat(),
            "bidCloseDate": "2026-06-10",
            "priceInit":   "15000000000",
            "bidForm":     "Đấu thầu rộng rãi",
            "bidMode":     "Trong nước",
            "status":      "01",
            "analysis":    "Product: MacBook Pro M4\nValue: 15.00B VND\nDeadline: 19 days left — plenty of time\nPriority: High — large MacBook procurement, ministry-level buyer",
            "crawled_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
            "source_url":  "https://muasamcong.mpi.gov.vn/web/guest/contractor-selection?render=detail&notifyId=TEST-001",
        },
        {
            "notifyId":    "TEST-002",
            "keyword":     "ipad",
            "notifyNo":    "TEST20260522-B",
            "bid_name":    "[TEST] Trang bị iPad cho phòng học thông minh THPT Chu Văn An",
            "investorName": "Sở Giáo dục và Đào tạo TP.HCM",
            "investorCode": "vn0000000002",
            "prov_name":   "Thành phố Hồ Chí Minh",
            "publicDate":  date.today().isoformat(),
            "bidCloseDate": "2026-05-27",
            "priceInit":   "2500000000",
            "bidForm":     "Chào hàng cạnh tranh",
            "bidMode":     "Trong nước",
            "status":      "01",
            "analysis":    "Product: iPad\nValue: 2.50B VND\nDeadline: 5 days left — urgent (!)\nPriority: High — direct iPad bid, very tight deadline",
            "crawled_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
            "source_url":  "https://muasamcong.mpi.gov.vn/web/guest/contractor-selection?render=detail&notifyId=TEST-002",
        },
        {
            "notifyId":    "TEST-003",
            "keyword":     "laptop",
            "notifyNo":    "TEST20260522-C",
            "bid_name":    "[TEST] Mua sắm laptop văn phòng Sở Tài chính Đà Nẵng",
            "investorName": "Sở Tài chính Đà Nẵng",
            "investorCode": "vn0000000003",
            "prov_name":   "Đà Nẵng",
            "publicDate":  date.today().isoformat(),
            "bidCloseDate": "2026-06-20",
            "priceInit":   "800000000",
            "bidForm":     "Chào hàng cạnh tranh",
            "bidMode":     "Trong nước",
            "status":      "01",
            "analysis":    "Product: Laptop (generic)\nValue: 0.80B VND\nDeadline: 29 days left — plenty of time\nPriority: Low — generic laptop bid, no Apple mention",
            "crawled_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
            "source_url":  "https://muasamcong.mpi.gov.vn/web/guest/contractor-selection?render=detail&notifyId=TEST-003",
        },
    ]
    print("Sending test email → minh_dao@apple.com only...")
    send_email(dummy, recipients=["minh_dao@apple.com"])
    print("Done — check minh_dao@apple.com inbox.")


def send_only():
    """Read today's crawled records from sheet and send email (no crawling)."""
    print(f"=== Apple Monitor (send-only) — {datetime.now().strftime('%Y-%m-%d %H:%M')} ===\n")

    print("Fetching full sheet for Excel export...")
    ws = connect_sheet()
    all_rows = ws.get_all_records(head=1)
    all_sheet = [{col: str(row.get(col, "")) for col in SHEET_COLS} for row in all_rows]
    print(f"  {len(all_sheet):,} total records in sheet")

    today = date.today().isoformat()
    crawled_today = [r for r in all_sheet if r.get("crawled_at", "").startswith(today)]

    # Filter to Apple-relevant device bids only — email is for Apple team action
    # Category crawl (HH+HON_HOP) ingests all goods; email shows only what a partner can propose Apple for
    _APPLE_DEV_KW = [
        "máy tính xách tay", "laptop", "máy vi tính xách tay",
        "máy tính bảng", "tablet", "máy vi tính bảng",
        "máy tính để bàn", "desktop", "máy vi tính để bàn", "all-in-one",
        "điện thoại thông minh", "smartphone", "điện thoại di động",
        "apple", "iphone", "ipad", "macbook", "imac", "mac pro",
        "mac mini", "mac studio", "airpods",
        "máy tính cá nhân", "máy vi tính",
    ]
    def _is_apple_relevant(bid_name):
        n = bid_name.lower()
        return any(k in n for k in _APPLE_DEV_KW)

    all_new = [r for r in crawled_today if _is_apple_relevant(r.get("bid_name", ""))]
    skipped = len(crawled_today) - len(all_new)
    print(f"  {len(crawled_today)} crawled today → {len(all_new)} Apple-relevant (skipped {skipped} non-device)")

    if not all_new:
        print("No records crawled today — no email sent.")
        return

    print("Sending email...")
    send_email(all_new, all_sheet)
    print(f"\n✓ Done — email sent with {len(all_new)} new records.")


if __name__ == "__main__":
    cmd = sys.argv[1].lower() if len(sys.argv) > 1 else "run"
    if cmd == "test":
        run_test()
    elif cmd == "crawl":
        run(do_send_email=False)
    elif cmd == "send-only":
        send_only()
    else:
        run()

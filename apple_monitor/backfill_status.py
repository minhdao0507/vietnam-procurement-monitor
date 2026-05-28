"""
backfill_status.py
Two-pass fix for stale status values across all Sheet records.

Pass 1 (no API): records that already have winner data but show Open status
                 → set status = PUB_KQLCNT directly.
Pass 2 (API):    closed bids with stale status and no winner
                 → re-fetch status + winner + price from API.
"""
import time
import random
from datetime import date
from apple_monitor import (
    connect_sheet, make_session, SHEET_COLS,
    BASE_URL, API_PATH, HEADERS, _STALE_STATUSES,
)
from apple_monitor_config import API_TOKEN
from openpyxl.utils import get_column_letter

BATCH = 100


def run():
    print("Connecting to sheet...")
    ws   = connect_sheet()
    rows = ws.get_all_records(head=1)
    today = str(date.today())

    sc = get_column_letter(SHEET_COLS.index("status") + 1)
    wc = get_column_letter(SHEET_COLS.index("winner") + 1)
    pc = get_column_letter(SHEET_COLS.index("winner_price") + 1)

    # ── Pass 1: winner exists + stale status → PUB_KQLCNT (no API calls) ──
    pass1 = [
        (i + 2, r)
        for i, r in enumerate(rows)
        if r.get("winner", "").strip()
        and r.get("status", "") in _STALE_STATUSES
    ]
    print(f"\nPass 1 (no API): {len(pass1)} records with winner + stale status → PUB_KQLCNT")
    if pass1:
        updates = [{"range": f"{sc}{row}", "values": [["PUB_KQLCNT"]]} for row, _ in pass1]
        for i in range(0, len(updates), BATCH):
            ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates[i:i + BATCH]})
            time.sleep(2)
        print(f"  Done — {len(pass1)} updated")

    # ── Pass 2: closed + stale status + no winner → re-fetch from API ──
    pass2 = [
        (i + 2, r)
        for i, r in enumerate(rows)
        if r.get("bidCloseDate", "") < today
        and r.get("status", "") in _STALE_STATUSES
        and not r.get("winner", "").strip()
        and r.get("notifyNo", "").strip()
    ]
    print(f"\nPass 2 (API): {len(pass2)} closed bids without winner → re-fetch status")
    if not pass2:
        print("  Nothing to do.")
    else:
        session = make_session()
        url     = f"{BASE_URL}{API_PATH}?token={API_TOKEN}"
        updates = []
        updated = 0
        skipped = 0

        for sheet_row, record in pass2:
            notify_no = record["notifyNo"]
            payload = [{
                "pageSize": 1, "pageNumber": 0,
                "query": [{
                    "index": "es-contractor-selection",
                    "keyWord": notify_no,
                    "matchType": "all-1",
                    "matchFields": ["notifyNo"],
                    "filters": [{"fieldName": "type", "searchType": "in",
                                 "fieldValues": ["es-notify-contractor"]}],
                }],
            }]
            try:
                resp  = session.post(url, headers=HEADERS, json=payload, verify=False, timeout=15)
                data  = resp.json()
                items = (data[0] if isinstance(data, list) else data).get("page", {}).get("content", [])
                if not items:
                    skipped += 1
                    continue
                it         = items[0]
                new_status = it.get("status", "")
                if not new_status:
                    skipped += 1
                    continue

                winner_raw = it.get("winningContractorName", "")
                winner     = (" | ".join(str(v) for v in winner_raw if v)
                              if isinstance(winner_raw, list) else str(winner_raw or ""))
                price_raw  = it.get("bidWinningPrice", "")
                price      = (str(price_raw[0]) if isinstance(price_raw, list) and price_raw
                              else str(price_raw or ""))

                updates.append({"range": f"{sc}{sheet_row}", "values": [[new_status]]})
                if winner:
                    updates.append({"range": f"{wc}{sheet_row}", "values": [[winner]]})
                    updates.append({"range": f"{pc}{sheet_row}", "values": [[price]]})
                updated += 1
            except Exception as e:
                print(f"  [err] {notify_no}: {e}")
                skipped += 1

            if len(updates) >= BATCH * 3:
                ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates})
                print(f"  Progress: {updated} updated, {skipped} no data")
                updates = []

            time.sleep(random.uniform(0.3, 0.6))

        if updates:
            ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates})

        print(f"  Done — {updated} updated, {skipped} no data found")

    print(f"\nBackfill complete. Pass 1: {len(pass1)} | Pass 2 API calls done.")


if __name__ == "__main__":
    run()

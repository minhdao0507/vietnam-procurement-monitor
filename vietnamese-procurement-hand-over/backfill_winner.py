"""
backfill_winner.py
Fetches winner + price for closed bids by searching each notifyNo individually.
Only processes bids where bidCloseDate < today and winner is empty.
Commits every BATCH records.
"""
import time
import random
from datetime import date
from apple_monitor import (
    connect_sheet, make_session, SHEET_COLS,
    BASE_URL, API_PATH, HEADERS,
)
from apple_monitor_config import API_TOKEN

BATCH = 50


def fetch_winner(session, notify_no):
    url = f"{BASE_URL}{API_PATH}?token={API_TOKEN}"
    payload = [{
        "pageSize": 1, "pageNumber": 0,
        "query": [{
            "index": "es-contractor-selection",
            "keyWord": notify_no,
            "matchType": "all-1",
            "matchFields": ["notifyNo"],
            "filters": [{"fieldName": "type", "searchType": "in", "fieldValues": ["es-notify-contractor"]}],
        }],
    }]
    try:
        resp = session.post(url, headers=HEADERS, json=payload, verify=False, timeout=15)
        data = resp.json()
        items = (data[0] if isinstance(data, list) else data).get("page", {}).get("content", [])
        if not items:
            return None, None
        it = items[0]
        winner_raw = it.get("winningContractorName", "")
        winner = " | ".join(str(v) for v in winner_raw if v) if isinstance(winner_raw, list) else str(winner_raw or "")
        price_raw = it.get("bidWinningPrice", "")
        price = str(price_raw[0]) if isinstance(price_raw, list) and price_raw else str(price_raw or "")
        return winner or None, price or None
    except Exception as e:
        print(f"  [err] {notify_no}: {e}")
        return None, None


def run():
    print("Connecting to sheet...")
    ws = connect_sheet()
    rows = ws.get_all_records(head=1)
    today = str(date.today())

    to_fix = [
        (i + 2, r)
        for i, r in enumerate(rows)
        if r.get("bidCloseDate", "") < today
        and r.get("notifyNo", "").strip()
        and not r.get("winner", "").strip()
    ]
    print(f"  {len(to_fix):,} closed bids without winner data")
    if not to_fix:
        print("  Nothing to fix.")
        return

    winner_col = SHEET_COLS.index("winner") + 1
    price_col  = SHEET_COLS.index("winner_price") + 1
    from openpyxl.utils import get_column_letter
    wc = get_column_letter(winner_col)
    pc = get_column_letter(price_col)

    session  = make_session()
    updates  = []
    updated  = 0
    skipped  = 0

    for sheet_row, record in to_fix:
        notify_no = record["notifyNo"]
        winner, price = fetch_winner(session, notify_no)

        if winner:
            updates.append({"range": f"{wc}{sheet_row}", "values": [[winner]]})
            updates.append({"range": f"{pc}{sheet_row}", "values": [[price or ""]]})
            updated += 1
        else:
            skipped += 1

        if len(updates) >= BATCH * 2:
            ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates})
            print(f"  Progress: {updated} updated, {skipped} no winner data")
            updates = []

        time.sleep(random.uniform(0.3, 0.6))

    if updates:
        ws.spreadsheet.values_batch_update({"valueInputOption": "RAW", "data": updates})

    print(f"\nDone. Updated: {updated} | No winner data: {skipped}")


if __name__ == "__main__":
    run()
